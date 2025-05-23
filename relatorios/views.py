import re
import xml.etree.ElementTree as ET
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import json
from datetime import date
from django.db.models import Count, Min, Max, Avg

from paciente.models import Paciente, Doenca as DoencaPaciente, Medicamento as MedicamentoPaciente, Anamnese
from consulta.models import Consulta, ProblemaSaude, PlanoAtuacao, Avaliacao, Medicamento as MedicamentoConsulta

# --------------------------------------------------------------------------------------
# --- Funções Utilitárias ---
# --------------------------------------------------------------------------------------

def limpar_nome_aba(nome):
    nome_limpo = re.sub(r'[\\/*?:\[\]]', '', nome)
    return nome_limpo[:31]

# --------------------------------------------------------------------------------------
# --- Funções de Exportação ---
# --------------------------------------------------------------------------------------

def exportar_dashboard(request):
    exportar = request.GET.getlist('exportar')
    formato = request.GET.get('formato', 'xml')

    if formato == 'xml':
        return exportar_xml(exportar)
    elif formato == 'xlsx':
        return exportar_xlsx(exportar)
    else:
        return HttpResponse('Formato não suportado', status=400)

def exportar_xml(exportar):
    root = ET.Element('Dashboard')

    if 'pacientes' in exportar:
        pacientes_tag = ET.SubElement(root, 'Pacientes')
        pacientes = Paciente.objects.all()
        for p in pacientes:
            paciente = ET.SubElement(pacientes_tag, 'Paciente')
            ET.SubElement(paciente, 'Nome').text = p.nome
            ET.SubElement(paciente, 'Nascimento').text = str(p.data_nascimento)
            ET.SubElement(paciente, 'Telefone').text = p.telefone or "Não Informado"
            ET.SubElement(paciente, 'Bairro').text = p.bairro or "Não Informado"
            ET.SubElement(paciente, 'Municipio').text = p.municipio or "Não Informado"

    if 'consultas' in exportar:
        consultas_tag = ET.SubElement(root, 'Consultas')
        consultas = Consulta.objects.all()
        for c in consultas:
            consulta = ET.SubElement(consultas_tag, 'Consulta')
            ET.SubElement(consulta, 'Paciente').text = c.paciente.nome
            ET.SubElement(consulta, 'DataConsulta').text = str(c.data_consulta)
            ET.SubElement(consulta, 'MotivoConsulta').text = c.motivo_consulta or "Não informado"
            ET.SubElement(consulta, 'Evolucao').text = c.evolucao or "Não informado"

    if 'intervencoes' in exportar:
        intervencoes_tag = ET.SubElement(root, 'Intervencoes')
        planos = PlanoAtuacao.objects.exclude(classificacao_intervencao__isnull=True)
        for p in planos:
            intervencao = ET.SubElement(intervencoes_tag, 'Intervencao')
            ET.SubElement(intervencao, 'Paciente').text = p.consulta.paciente.nome
            ET.SubElement(intervencao, 'DataConsulta').text = str(p.consulta.data_consulta)
            ET.SubElement(intervencao, 'Classificacao').text = p.classificacao_intervencao or "Não informado"
            ET.SubElement(intervencao, 'Resultado').text = p.resultado or "Não informado"

    if 'rnms' in exportar:
        rnms_tag = ET.SubElement(root, 'RNMs')
        avaliacoes = Avaliacao.objects.select_related('medicamento__consulta__paciente').all()
        for a in avaliacoes:
            rnm = ET.SubElement(rnms_tag, 'RNM')
            ET.SubElement(rnm, 'Paciente').text = a.medicamento.consulta.paciente.nome
            ET.SubElement(rnm, 'Medicamento').text = a.medicamento.nome
            ET.SubElement(rnm, 'Necessidade').text = str(a.necessidade) if a.necessidade is not None else "Não informado"
            ET.SubElement(rnm, 'Seguranca').text = str(a.seguranca) if a.seguranca is not None else "Não informado"
            ET.SubElement(rnm, 'Efetividade').text = str(a.efetividade) if a.efetividade is not None else "Não informado"


    response = HttpResponse(content_type='application/xml')
    response['Content-Disposition'] = 'attachment; filename="dashboard_farmacia.xml"'
    tree = ET.ElementTree(root)
    tree.write(response, encoding='unicode')
    return response

def exportar_xlsx(exportar):
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

    if 'pacientes' in exportar:
        ws = wb.create_sheet(limpar_nome_aba('Pacientes'))
        ws.append(['Nome', 'Nascimento', 'Telefone', 'Bairro', 'Municipio'])
        style_header(ws)
        pacientes = Paciente.objects.all()
        for p in pacientes:
            ws.append([p.nome, str(p.data_nascimento), p.telefone or "Não Informado", p.bairro or "Não Informado", p.municipio or "Não Informado"])

    if 'consultas' in exportar:
        ws = wb.create_sheet(limpar_nome_aba('Consultas'))
        ws.append(['Paciente', 'DataConsulta', 'MotivoConsulta', 'Evolucao'])
        style_header(ws)
        consultas = Consulta.objects.all()
        for c in consultas:
            ws.append([c.paciente.nome, str(c.data_consulta), c.motivo_consulta or "Não informado", c.evolucao or "Não informado"])

    if 'intervencoes' in exportar:
        ws = wb.create_sheet(limpar_nome_aba('Intervencoes'))
        ws.append(['Paciente', 'DataConsulta', 'Classificacao', 'Resultado'])
        style_header(ws)
        planos = PlanoAtuacao.objects.exclude(classificacao_intervencao__isnull=True)
        for p in planos:
            ws.append([p.consulta.paciente.nome, str(p.consulta.data_consulta), p.classificacao_intervencao or "Não informado", p.resultado or "Não informado"])

    if 'rnms' in exportar:
        ws = wb.create_sheet(limpar_nome_aba('RNMs'))
        ws.append(['Paciente', 'Medicamento', 'Necessidade', 'Seguranca', 'Efetividade'])
        style_header(ws)
        avaliacoes = Avaliacao.objects.select_related('medicamento__consulta__paciente').all()
        for a in avaliacoes:
            ws.append([
                a.medicamento.consulta.paciente.nome,
                a.medicamento.nome,
                str(a.necessidade) if a.necessidade is not None else "Não informado",
                str(a.seguranca) if a.seguranca is not None else "Não informado",
                str(a.efetividade) if a.efetividade is not None else "Não informado"
            ])


    if 'anamneses' in exportar:
        ws = wb.create_sheet(limpar_nome_aba('Anamneses'))
        ws.append(['Paciente', 'Data', 'Esquecimentos', 'Percepção de Saúde'])
        style_header(ws)
        anamneses = Anamnese.objects.select_related('paciente').all()
        for a in anamneses:
            ws.append([
                a.paciente.nome,
                str(getattr(a, 'data', 'Não informado')),
                a.autonomia_medicamentos.esquecimentos if a.autonomia_medicamentos else 'Não informado',
                a.perfil_clinico.percepcao_saude if a.perfil_clinico else 'Não informado',
            ])

    if not wb.sheetnames:
        ws = wb.create_sheet('Sem Dados')
        ws.append(['Nenhum dado selecionado'])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="dashboard_farmacia.xlsx"'
    wb.save(response)
    return response

# --------------------------------------------------------------------------------------
# --- Funções Auxiliares de Análise ---
# --------------------------------------------------------------------------------------

def aplicar_filtros(pacientes, consultas, data_inicio=None, data_fim=None,
                    doenca_paciente_id=None, medicamento_paciente_id=None,
                    faixa_etaria_filtro=None):
    if doenca_paciente_id:
        pacientes = pacientes.filter(doencas__id=doenca_paciente_id)
    if medicamento_paciente_id:
        pacientes = pacientes.filter(medicamentos__id=medicamento_paciente_id)
    if faixa_etaria_filtro:
        try:
            idade_min, idade_max = map(int, faixa_etaria_filtro.split('-'))
            hoje = date.today()
            data_max = hoje.replace(year=hoje.year - idade_min)
            data_min = hoje.replace(year=hoje.year - idade_max - 1)
            pacientes = pacientes.filter(data_nascimento__range=(data_min, data_max))
        except ValueError:
            pass

    if data_inicio and data_fim:
        consultas = consultas.filter(data_consulta__range=[data_inicio, data_fim])

    consultas = consultas.filter(paciente__in=pacientes)
    planos = PlanoAtuacao.objects.filter(consulta__in=consultas)

    return pacientes, consultas, planos

def calcular_estatisticas(pacientes, consultas, planos, medicamentos_consulta, problemas_consulta):
    total_pacientes = pacientes.count()
    total_consultas = consultas.count()
    total_intervencoes = planos.exclude(classificacao_intervencao__isnull=True).count()
    
    total_rnm = Avaliacao.objects.filter(medicamento__consulta__in=consultas).count()
    rnm_resolvido = planos.filter(rnm_resolvido=True).count()

    media_intervencoes = planos.values('consulta__paciente').annotate(total=Count('id')).aggregate(avg=Avg('total'))['avg'] or 0
    media_medicamentos = medicamentos_consulta.values('consulta__paciente').annotate(total=Count('id')).aggregate(avg=Avg('total'))['avg'] or 0
    media_rnm = Avaliacao.objects.filter(medicamento__consulta__in=consultas).values('medicamento__consulta__paciente').annotate(total=Count('id')).aggregate(avg=Avg('total'))['avg'] or 0
    media_consultas_paciente = consultas.values('paciente').annotate(total=Count('id')).aggregate(avg=Avg('total'))['avg'] or 0

    esquecem = Anamnese.objects.filter(paciente__in=pacientes, autonomia__esquecimentos='sim').count()

    return {
        'total_pacientes': total_pacientes,
        'total_consultas': total_consultas,
        'total_intervencoes': total_intervencoes,
        'total_rnm': total_rnm,
        'rnm_resolvido': rnm_resolvido,
        'media_intervencoes': round(media_intervencoes, 1),
        'media_medicamentos': round(media_medicamentos, 1),
        'media_rnm': round(media_rnm, 1),
        'media_consultas_paciente': round(media_consultas_paciente, 1),
        'esquecem': esquecem
    }


def calcular_distribuicoes(pacientes, consultas, planos):
    grupos = [(0,18), (19,40), (41,60), (61,120)]
    labels_faixa = ["0-18", "19-40", "41-60", "60+"]
    valores_faixa = []
    hoje = date.today()
    for min_age, max_age in grupos:
        data_max = hoje.replace(year=hoje.year - min_age)
        data_min = hoje.replace(year=hoje.year - max_age - 1)
        valores_faixa.append(pacientes.filter(data_nascimento__range=(data_min, data_max)).count())

    genero_data = pacientes.values('genero').annotate(total=Count('id'))
    labels_genero = [g['genero'] or 'Não informado' for g in genero_data]
    valores_genero = [g['total'] for g in genero_data]

    intervencao_data = planos.values('classificacao_intervencao').annotate(total=Count('id'))
    labels_inter = [i['classificacao_intervencao'] or 'Não informado' for i in intervencao_data]
    valores_inter = [i['total'] for i in intervencao_data]

    return {
        'labels_faixa': json.dumps(labels_faixa),
        'valores_faixa': json.dumps(valores_faixa),
        'labels_genero': json.dumps(labels_genero),
        'valores_genero': json.dumps(valores_genero),
        'labels_inter': json.dumps(labels_inter),
        'valores_inter': json.dumps(valores_inter),
    }

def calcular_top10_por_consultas(model_qs, campo):
    qs = model_qs.annotate(total=Count('consulta')).order_by('-total')[:10]
    labels = [getattr(obj, campo) for obj in qs]
    valores = [obj.total for obj in qs]
    return json.dumps(labels), json.dumps(valores)

def calcular_top10_por_pacientes(model_qs, campo):
    qs = model_qs.annotate(total=Count('pacientes')).order_by('-total')[:10]
    labels = [getattr(obj, campo) for obj in qs]
    valores = [obj.total for obj in qs]
    return json.dumps(labels), json.dumps(valores)



def calcular_armazenamento_descarte(pacientes):
    armazenamento = Anamnese.objects.filter(paciente__in=pacientes).values('autonomia__local_guarda').annotate(total=Count('id'))
    descarte = Anamnese.objects.filter(paciente__in=pacientes).values('autonomia__forma_descarte').annotate(total=Count('id'))
    return armazenamento, descarte

def calcular_intervalo_consultas():
    return Consulta.objects.values('paciente__nome').annotate(
        primeira=Min('data_consulta'),
        ultima=Max('data_consulta'),
        numero_consultas=Count('id')
    )

def calcular_percepcao_saude(pacientes):
    return Anamnese.objects.filter(
        paciente__in=pacientes, saude__percepcao_saude__isnull=False
    ).values('saude__percepcao_saude').annotate(
        total=Count('id')
    ).order_by('saude__percepcao_saude')

# --------------------------------------------------------------------------------------
# --- View Principal ---
# --------------------------------------------------------------------------------------

def dashboard_clinico(request):
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    doenca_paciente_id = request.GET.get('doenca_paciente')
    medicamento_paciente_id = request.GET.get('medicamento_paciente')
    faixa_etaria_filtro = request.GET.get('faixa_etaria')
    classe_medicamento_filtro = request.GET.get('classe_medicamento')
    problema_saude_filtro = request.GET.get('problema_saude')

    pacientes = Paciente.objects.all()
    consultas = Consulta.objects.all()

    pacientes, consultas, planos = aplicar_filtros(
        pacientes, consultas, data_inicio, data_fim,
        doenca_paciente_id, medicamento_paciente_id,
        faixa_etaria_filtro
    )

    problemas_consulta = ProblemaSaude.objects.filter(consulta__in=consultas)
    medicamentos_consulta = MedicamentoConsulta.objects.filter(consulta__in=consultas)

    if classe_medicamento_filtro:
        medicamentos_consulta = medicamentos_consulta.filter(classe__icontains=classe_medicamento_filtro)
    if problema_saude_filtro:
        problemas_consulta = problemas_consulta.filter(problema__icontains=problema_saude_filtro)

    estatisticas = calcular_estatisticas(pacientes, consultas, planos, medicamentos_consulta, problemas_consulta)
    distribuicoes = calcular_distribuicoes(pacientes, consultas, planos)

    labels_problema_consulta, valores_problema_consulta = calcular_top10_por_consultas(
        ProblemaSaude.objects.filter(consulta__in=consultas),
        'problema'
    )

    labels_medicamento_consulta, valores_medicamento_consulta = calcular_top10_por_consultas(
        MedicamentoConsulta.objects.filter(consulta__in=consultas),
        'nome'
    )



    armazenamento, descarte = calcular_armazenamento_descarte(pacientes)
    rnm_tipos = planos.values('registro_intervencao').annotate(total=Count('id'))
    intervalo_consultas = calcular_intervalo_consultas()
    percepcao = calcular_percepcao_saude(pacientes)
    taxa_aceitacao = calcular_taxa_aceitacao(planos)
    polimedicados = calcular_pacientes_polimedicados(consultas)
    adesao = calcular_adesao_seguimento(consultas)
    tempo_medio_intervalo = calcular_tempo_medio_entre_consultas(intervalo_consultas)
    labels_doenca_paciente, valores_doenca_paciente = calcular_top10_por_pacientes(
    DoencaPaciente.objects.filter(pacientes__in=pacientes),
    'nome'
    )

    labels_medicamento_paciente, valores_medicamento_paciente = calcular_top10_por_pacientes(
        MedicamentoPaciente.objects.filter(pacientes__in=pacientes),
        'nome'
    )

    context = {
        **estatisticas,
        **distribuicoes,
        'taxa_aceitacao': taxa_aceitacao,
        'polimedicados': polimedicados,
        'adesao': adesao,
        'tempo_medio_intervalo': tempo_medio_intervalo,
        
        'labels_doenca_paciente': labels_doenca_paciente,
        'valores_doenca_paciente': valores_doenca_paciente,
        'labels_medicamento_paciente': labels_medicamento_paciente,
        'valores_medicamento_paciente': valores_medicamento_paciente,

        'labels_problema_consulta': labels_problema_consulta,
        'valores_problema_consulta': valores_problema_consulta,
        'labels_medicamento_consulta': labels_medicamento_consulta,
        'valores_medicamento_consulta': valores_medicamento_consulta,
        
        'armazenamento': armazenamento,
        'descarte': descarte,
        'rnm_tipos': rnm_tipos,
        'intervalo_consultas': intervalo_consultas,
        'percepcao': percepcao,
        
        'filtros': {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'doenca_paciente': doenca_paciente_id,
            'medicamento_paciente': medicamento_paciente_id,
            'faixa_etaria_filtro': faixa_etaria_filtro,
            'classe_medicamento_filtro': classe_medicamento_filtro,
            'problema_saude_filtro': problema_saude_filtro,
        },
        'dropdowns': {
            'doencas': DoencaPaciente.objects.all(),
            'medicamentos': MedicamentoPaciente.objects.all(),
            'classes_medicamento': MedicamentoConsulta.objects.values_list('classe', flat=True).distinct(),
            'problemas_saude': ProblemaSaude.objects.values_list('problema', flat=True).distinct(),
        },
        'exportar_opcoes': [
            ("Pacientes Atendidos", "pacientes"),
            ("Consultas", "consultas"),
            ("Intervenções Realizadas", "intervencoes"),
            ('RNMs', 'rnms'),
            ("Anamneses", "anamneses"),
        ]
    }


    return render(request, 'relatorios/dashboard.html', context)

def calcular_taxa_aceitacao(planos):
    total_intervencoes = planos.count()
    if total_intervencoes == 0:
        return 0
    aceitas = planos.filter(resultado='Aceita').count()
    taxa = (aceitas / total_intervencoes) * 100
    return round(taxa, 1)

def calcular_pacientes_polimedicados(consultas, limite=5):
    medicamentos_consulta = MedicamentoConsulta.objects.filter(consulta__in=consultas)
    polimedicados = medicamentos_consulta.values('consulta__paciente').annotate(num=Count('id')).filter(num__gte=limite).count()
    return polimedicados

def calcular_adesao_seguimento(consultas):
    consultas_por_paciente = consultas.values('paciente').annotate(num=Count('id'))
    retornaram = consultas_por_paciente.filter(num__gt=1).count()
    nao_retornaram = consultas_por_paciente.filter(num=1).count()
    total = retornaram + nao_retornaram
    if total == 0:
        taxa_retorno = 0
    else:
        taxa_retorno = (retornaram / total) * 100
    return {
        'retornaram': retornaram,
        'nao_retornaram': nao_retornaram,
        'taxa_retorno': round(taxa_retorno, 1)
    }
def calcular_tempo_medio_entre_consultas(intervalo_consultas):
    intervalos = []
    for item in intervalo_consultas:
        primeira = item['primeira']
        ultima = item['ultima']
        num_consultas = item['numero_consultas']
        if num_consultas > 1:
            dias = (ultima - primeira).days
            if dias > 0:
                media = dias / (num_consultas - 1)
                intervalos.append(media)
    if not intervalos:
        return 0
    media_intervalo = sum(intervalos) / len(intervalos)
    return round(media_intervalo, 1)
